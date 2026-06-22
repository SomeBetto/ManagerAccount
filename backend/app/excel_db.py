import openpyxl
import os
import json

CONFIG_FILE = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config.json')

SCHEMA = {
    'accounts': ['id', 'email', 'password', 'pin'],
    'characters': ['id', 'account_id', 'name', 'password', 'level', 'class_name', 'char_type', 'is_favorite'],
    'items': ['id', 'character_id', 'name', 'item_type', 'description'],
    'level_entries': ['id', 'character_id', 'priority', 'note'],
    'daily_events': ['id', 'name', 'description', 'start_date', 'end_date'],
    'daily_event_participants': ['id', 'event_id', 'character_id'],
    'daily_event_progress': ['id', 'participant_id', 'event_date', 'is_completed'],
    'coupons': ['id', 'code', 'name', 'description'],
    'coupon_redemptions': ['id', 'coupon_id', 'account_id', 'character_id', 'date']
}

HEADER_ALIASES = {
    'id': ['id', 'no', '#', 'numero', 'nº', 'id_character', 'id_cuenta', 'cuenta_id'],
    'email': ['email', 'correo', 'cuenta', 'account', 'acc', 'usuario', 'user'],
    'password': ['password', 'contraseña', 'contrasena', 'contrasea', 'pass', 'clave', 'pw'],
    'pin': ['pin', 'segunda clave', 'bank pin', 'pin2'],
    'account_id': ['account_id', 'id_cuenta', 'cuenta_id', 'acc_id', 'account'],
    'name': ['name', 'nombre', 'personaje', 'char_name', 'char name', 'character', 'pjs', 'pj'],
    'level': ['level', 'nivel', 'lvl', 'lv'],
    'class_name': ['class_name', 'clase', 'class', 'job', 'profesion', 'profesion'],
    'char_type': ['char_type', 'tipo', 'type', 'rol', 'role', 'categoria'],
    'is_favorite': ['is_favorite', 'favorito', 'fav', 'estrella', 'destacado'],
    'description': ['description', 'descripcion', 'desc', 'detalle', 'notas'],
    'item_type': ['item_type', 'tipo_item', 'tipo item', 'tipo', 'categoria'],
    'character_id': ['character_id', 'id_personaje', 'char_id', 'pj_id'],
    'priority': ['priority', 'prioridad', 'orden', 'rango'],
    'note': ['note', 'nota', 'notas', 'observacion', 'obs'],
    'start_date': ['start_date', 'fecha_inicio', 'inicio', 'fecha inicio', 'desde'],
    'end_date': ['end_date', 'fecha_fin', 'fin', 'fecha fin', 'cierre', 'hasta'],
    'event_date': ['event_date', 'fecha', 'dia', 'date', 'momento'],
    'is_completed': ['is_completed', 'completado', 'listo', 'done', 'terminado'],
    'code': ['code', 'codigo', 'cod', 'coupon_code', 'cupon'],
    'coupon_id': ['coupon_id', 'id_cupon', 'cupon_id', 'ref_cupon'],
    'date': ['date', 'fecha', 'creado', 'dia', 'timestamp']
}

def get_excel_path():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
            return data.get('excel_path')
    return None

def set_excel_path(path):
    data = {}
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
    data['excel_path'] = path
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=4)

def get_config_key(key):
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
            return data.get(key)
    return None

def set_config_keys(updates):
    data = {}
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
    for k, v in updates.items():
        data[k] = v
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=4)

class ExcelDB:
    @staticmethod
    def _get_workbook():
        path = get_excel_path()
        if not path or not os.path.exists(path):
            raise Exception("Ruta de archivo Excel no configurada o el archivo no existe.")
        # data_only=True is CRITICAL to read evaluated VLOOKUPs instead of strings
        return openpyxl.load_workbook(path, data_only=True), path

    @staticmethod
    def _save_workbook(wb, path):
        wb.save(path)

    @staticmethod
    def _get_actual_sheetname(wb, target_name):
        target_lower = target_name.lower()
        for name in wb.sheetnames:
            if name.lower() == target_lower:
                return name
        return None

    @staticmethod
    def _map_header(raw_header):
        if not raw_header:
            return ""
        clean_header = str(raw_header).lower().strip()
        for canonical, aliases in HEADER_ALIASES.items():
            if clean_header in aliases:
                return canonical
        return clean_header

    @classmethod
    def init_db(cls):
        path = get_excel_path()
        if not path:
            return # Nothing to init if no path
        
        # If file doesn't exist, we could theoretically create it, but the instruction implies picking an existing one.
        if not os.path.exists(path):
            wb = openpyxl.Workbook()
            # Do not remove default sheet here to prevent "At least one sheet must be visible" error from openpyxl
            # Save it so we can load it
            cls._save_workbook(wb, path)
        
        wb, _ = cls._get_workbook()
        changed = False

        for table_name, columns in SCHEMA.items():
            actual_name = cls._get_actual_sheetname(wb, table_name)
            if not actual_name:
                ws = wb.create_sheet(title=table_name)
                # Write headers
                for col_num, column_title in enumerate(columns, 1):
                    ws.cell(row=1, column=col_num, value=column_title)
                changed = True
            else:
                # Check for missing columns in existing sheet
                ws = wb[actual_name]
                rows = list(ws.iter_rows(max_row=1))
                if rows:
                    existing_headers = [cls._map_header(cell.value) for cell in rows[0]]
                    next_col = len(rows[0]) + 1
                    for col_name in columns:
                        if col_name not in existing_headers:
                            # Add missing column
                            ws.cell(row=1, column=next_col, value=col_name.capitalize())
                            next_col += 1
                            changed = True

        # Remove the default sheet only after we have created at least one actual data sheet
        if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb['Sheet']
            changed = True

        if changed:
            cls._save_workbook(wb, path)

    @staticmethod
    def _row_to_dict(headers, row):
        return {headers[i]: cell.value for i, cell in enumerate(row)}

    @classmethod
    def get_all(cls, table_name):
        wb, _ = cls._get_workbook()
        actual_name = cls._get_actual_sheetname(wb, table_name)
        if not actual_name:
            return []
        ws = wb[actual_name]
        
        results = []
        rows = list(ws.iter_rows())
        if not rows:
            return []
            
        # DYNAMIC HEADER DETECTION
        # Scan first 10 rows to find the most likely header row
        header_row_idx = 0
        headers = []
        found_header_match = False
        
        for i, row in enumerate(rows[:10]):
            # Get cleaned values for checking
            raw_vals = [str(cell.value).lower().strip() if cell.value is not None else "" for cell in row]
            
            # Count matches against known header aliases
            known_matches = 0
            for val in raw_vals:
                if not val: continue
                # check if val is in ANY of the alias lists
                if any(val in aliases for aliases in HEADER_ALIASES.values()):
                    known_matches += 1
            
            if known_matches >= 2: # Found a row where at least 2 headers match known terms
                header_row_idx = i
                headers = [cls._map_header(cell.value) if cell.value is not None else f"col_{j}" 
                           for j, cell in enumerate(row)]
                found_header_match = True
                break
        
        if not found_header_match:
            # Fallback to the very first row
            headers = [cls._map_header(cell.value) if cell.value is not None else f"col_{j}" 
                       for j, cell in enumerate(rows[0])]
            header_row_idx = 0
        
        # Start reading data from the row AFTER the identified header
        for relative_idx, row in enumerate(rows[header_row_idx+1:], start=1):
            # Skip completely empty rows
            if all(cell.value is None for cell in row):
                continue
            
            data = cls._row_to_dict(headers, row)
            
            # Skip if row is effectively empty (only whitespace or None)
            is_row_empty = True
            for val in data.values():
                if val is not None and str(val).strip() != "":
                    is_row_empty = False
                    break
            if is_row_empty:
                continue
            
            # Physical ID resolution
            if 'id' in data and data['id'] is not None:
                try:
                    data['id'] = int(data['id'])
                except:
                    pass
            else:
                # Fallback to row index if ID is missing
                data['id'] = header_row_idx + relative_idx + 1
                    
            results.append(data)
            
        # Relational Bridging: Linking Characters to Accounts
        if table_name == 'characters':
            accounts_data = cls.get_all('accounts')
            # build map from email -> id
            email_to_id = {}
            for a in accounts_data:
                e = str(a.get('email', '') or '').lower().strip()
                if e:
                    email_to_id[e] = a.get('id')
            
            for data in results:
                # Link via the 'email' column (mapped from "CUENTA" or "Cuenta")
                c_email = str(data.get('email', '') or '').lower().strip()
                if c_email in email_to_id:
                    data['account_id'] = email_to_id[c_email]
                elif not data.get('account_id'):
                    data['account_id'] = None 
        
        return results

    @classmethod
    def get_by_id(cls, table_name, record_id):
        all_records = cls.get_all(table_name)
        for r in all_records:
            if r.get('id') == record_id:
                return r
        return None

    @classmethod
    def insert(cls, table_name, data_dict):
        wb, path = cls._get_workbook()
        actual_name = cls._get_actual_sheetname(wb, table_name)
        if not actual_name:
            ws = wb.create_sheet(title=table_name)
        else:
            ws = wb[actual_name]

        # 1. DYNAMIC HEADER DETECTION (Same as get_all)
        rows = list(ws.iter_rows())
        header_row_idx = 0
        headers = []
        found_header_match = False
        
        if rows:
            for i, row in enumerate(rows[:10]):
                raw_vals = [str(cell.value).lower().strip() if cell.value is not None else "" for cell in row]
                known_matches = sum(1 for v in raw_vals if v and any(v in aliases for aliases in HEADER_ALIASES.values()))
                
                if known_matches >= 2:
                    header_row_idx = i
                    headers = [cls._map_header(cell.value) if cell.value is not None else f"col_{j}" for j, cell in enumerate(row)]
                    found_header_match = True
                    break
        
        if not found_header_match:
            headers = SCHEMA.get(table_name, [])
            header_row_idx = 0
            if not rows:
                for col_num, title in enumerate(headers, 1):
                    ws.cell(row=1, column=col_num, value=title)

        # 2. RELATIONAL LOOKUP (Account ID -> Email)
        if table_name == 'characters' and 'account_id' in data_dict:
            # If account_id is an Int (numeric ID from frontend), we find its Email
            # to store in the "CUENTA" column (mapped to 'email').
            try:
                acc_id = int(data_dict['account_id'])
                acc = cls.get_by_id('accounts', acc_id)
                if acc:
                    data_dict['email'] = acc.get('email')
                else:
                    # Account not found by physical ID
                    pass
            except (ValueError, TypeError):
                # account_id is already an email or invalid
                pass
        
        # 3. GET CURRENT DATA & ADD NEW
        # We reuse get_all to get a clean list of existing data
        current_data = cls.get_all(table_name)
        
        # Generate new ID if not present (Physical ID lookup)
        all_ids = [d.get('id') for d in current_data if isinstance(d.get('id'), int)]
        new_id = max(all_ids) + 1 if all_ids else 1
        data_dict['id'] = new_id
        
        current_data.append(data_dict)

        # 4. SORTING
        if table_name == 'characters':
            # Sort by level descending, handle None/invalid levels
            def get_level(c):
                try: return int(c.get('level', 0) or 0)
                except: return 0
            current_data.sort(key=get_level, reverse=True)

        # 5. REWRITE SHEET
        # Delete all data rows after header
        # ws.delete_rows starts from header_row_idx + 2 (first data row)
        # We delete a large enough range to clear ghost rows like 1003
        max_to_delete = max(len(rows), 2000) # Ensure we cover row 1003
        if len(rows) > header_row_idx + 1:
            ws.delete_rows(header_row_idx + 2, max_to_delete)

        # Write data
        for i, data in enumerate(current_data):
            row_num = header_row_idx + 2 + i
            for col_idx, h in enumerate(headers, 1):
                val = data.get(h)
                # Don't overwrite the ID if it's already in the first column and not mapped to 'id'
                # Actually, our headers map row[0] to 'id' usually.
                ws.cell(row=row_num, column=col_idx, value=val)

        cls._save_workbook(wb, path)
        return data_dict

    @classmethod
    def update(cls, table_name, record_id, new_data):
        wb, path = cls._get_workbook()
        actual_name = cls._get_actual_sheetname(wb, table_name)
        if not actual_name:
            return None

        # Relational Lookup (Account ID -> Email)
        if table_name == 'characters' and 'account_id' in new_data:
            try:
                acc_id = int(new_data['account_id'])
                acc = cls.get_by_id('accounts', acc_id)
                if acc:
                    new_data['email'] = acc.get('email')
            except (ValueError, TypeError):
                pass
        
        ws = wb[actual_name]
        rows = list(ws.iter_rows())
        header_row_idx = 0
        headers = []
        found_header_match = False
        
        if rows:
            for i, row in enumerate(rows[:10]):
                raw_vals = [str(cell.value).lower().strip() if cell.value is not None else "" for cell in row]
                known_matches = sum(1 for v in raw_vals if v and any(v in aliases for aliases in HEADER_ALIASES.values()))
                if known_matches >= 2:
                    header_row_idx = i
                    headers = [cls._map_header(cell.value) if cell.value is not None else f"col_{j}" for j, cell in enumerate(row)]
                    found_header_match = True
                    break
        
        if not found_header_match:
            headers = [cls._map_header(cell.value) if (rows and cell.value) else f"col_{j}" for j, cell in enumerate(rows[0] if rows else [])]
            header_row_idx = 0
            
        # Identify ID column
        id_col_idx = -1
        for idx, h in enumerate(headers):
            if h == 'id':
                id_col_idx = idx
                break

        updated_dict = None
        for rel_idx, row in enumerate(rows[header_row_idx+1:], start=1):
            # Virtual ID Logic: matching get_all
            current_row_idx = header_row_idx + rel_idx + 1
            row_id = None
            
            if id_col_idx != -1 and row[id_col_idx].value is not None:
                try:
                    row_id = int(row[id_col_idx].value)
                except:
                    row_id = current_row_idx
            else:
                row_id = current_row_idx
            
            if row_id == record_id:
                # Found it
                for col_idx, h in enumerate(headers, 1):
                    if h in new_data:
                        val = new_data[h]
                        # Convert to Excel boolean if appropriate
                        if h == 'is_favorite' and isinstance(val, bool):
                            pass # openpyxl handles bool
                            
                        ws.cell(row=current_row_idx, column=col_idx, value=val)
                cls._save_workbook(wb, path)
                
                # Re-fetch as dict
                updated_row = list(ws.iter_rows(min_row=current_row_idx, max_row=current_row_idx))[0]
                updated_dict = cls._row_to_dict(headers, updated_row)
                updated_dict['id'] = row_id
                break
                
        return updated_dict

    @classmethod
    def get_catalog(cls, sheet_name, column_name):
        wb, path = cls._get_workbook()
        # Find sheet by case-insensitive name
        actual_name = cls._get_actual_sheetname(wb, sheet_name)
        if not actual_name:
            return []
            
        ws = wb[actual_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
            
        # Find column index from the first 10 rows (header detection)
        col_idx = -1
        header_row_idx = 0
        
        for i, row in enumerate(rows[:10]):
            for j, cell in enumerate(row):
                if cell and str(cell).lower().strip() == column_name.lower().strip():
                    col_idx = j
                    header_row_idx = i
                    break
            if col_idx != -1: break
            
        if col_idx == -1:
            return []
            
        values = []
        for row in rows[header_row_idx+1:]:
            val = row[col_idx]
            if val is not None:
                s_val = str(val).strip()
                if s_val and s_val not in values:
                    values.append(s_val)
                    
        return sorted(values)
